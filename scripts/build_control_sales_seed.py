#!/usr/bin/env python3
"""Build the idempotent Control de Ventas seed from the normalized workbook."""

import json
import sys
from collections import Counter
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook


def clean(value):
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def cents(value):
    if value in (None, ""):
        return None
    try:
        return int((Decimal(str(value)) * 100).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    except (InvalidOperation, ValueError):
        return None


def rows(sheet):
    headers = [clean(cell.value) for cell in sheet[4]]
    for values in sheet.iter_rows(min_row=5, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        yield {headers[index]: values[index] for index in range(len(headers))}


def main():
    if len(sys.argv) != 3:
        raise SystemExit("usage: build_control_sales_seed.py INPUT.xlsx OUTPUT.json")
    source = Path(sys.argv[1]).resolve()
    output = Path(sys.argv[2]).resolve()
    workbook = load_workbook(source, read_only=True, data_only=True)

    groups = {clean(item["grupo_id"]): item for item in rows(workbook["Grupos"])}
    anomalies_by_entity = {}
    for item in rows(workbook["Anomalias"]):
        key = (clean(item["entidad"]).lower(), clean(item["entidad_id"]))
        anomalies_by_entity.setdefault(key, []).append({
            "id": clean(item["anomalia_id"]),
            "severity": clean(item["severidad"]),
            "type": clean(item["tipo"]),
            "description": clean(item["descripcion"]),
            "recommendedAction": clean(item["accion_recomendada"]),
            "sourceRow": clean(item["fila_origen"]),
        })

    details_by_order = {}
    for item in rows(workbook["Detalles"]):
        order_external_id = clean(item["orden_id"])
        group_external_id = clean(item["grupo_id"])
        group = groups.get(group_external_id)
        quantity = clean(item["cantidad"])
        unit_price_cents = cents(item["precio_unitario"])
        vat_cents = cents(item["iva"]) or 0
        original_cents = cents(item["importe_original"])
        calculated_cents = cents(item["importe_calculado"])
        review_required = unit_price_cents is None
        line_total_cents = calculated_cents if calculated_cents is not None else original_cents
        details_by_order.setdefault(order_external_id, []).append({
            "externalId": clean(item["detalle_id"]),
            "groupExternalId": group_external_id,
            "sequence": int(item["secuencia_detalle"] or 0),
            "product": clean(group["descripcion"]) if group else clean(item["descripcion"]),
            "size": clean(item["descripcion"]) if group else "",
            "quantity": quantity,
            "unitPriceCents": unit_price_cents,
            "vatCents": vat_cents,
            "lineTotalCents": line_total_cents or 0,
            "originalTotalCents": original_cents,
            "sourceRow": clean(item["fila_origen"]),
            "qualityStatus": clean(item["estado_calidad"]),
            "notes": clean(item["observaciones"]),
            "reviewRequired": review_required,
            "anomalies": anomalies_by_entity.get(("detalle", clean(item["detalle_id"])), []),
        })

    orders = []
    for item in rows(workbook["Ordenes"]):
        external_id = clean(item["orden_id"])
        details = details_by_order.get(external_id, [])
        order_anomalies = anomalies_by_entity.get(("orden", external_id), [])
        if clean(item["estado_calidad"]).lower() not in {"", "ok", "valido", "válido"}:
            order_anomalies = [*order_anomalies, {
                "id": f"quality-{external_id}",
                "severity": "Advertencia",
                "type": "estado_calidad",
                "description": clean(item["observaciones"]) or clean(item["estado_calidad"]),
                "recommendedAction": "Revisar el registro histórico antes de modificarlo.",
                "sourceRow": clean(item["fila_origen_inicio"]),
            }]
        orders.append({
            "externalId": external_id,
            "number": clean(item["numero_orden"]),
            "date": clean(item["fecha"]),
            "originalDate": clean(item["fecha_original"]),
            "seller": clean(item["vendedor"]),
            "originalSeller": clean(item["vendedor_original"]),
            "client": clean(item["cliente"]),
            "declaredTotalCents": cents(item["total_declarado"]),
            "sourceRowStart": clean(item["fila_origen_inicio"]),
            "sourceRowEnd": clean(item["fila_origen_fin"]),
            "qualityStatus": clean(item["estado_calidad"]),
            "notes": clean(item["observaciones"]),
            "anomalies": order_anomalies,
            "details": details,
        })

    number_counts = Counter(order["number"].casefold() for order in orders if order["number"])
    for order in orders:
        if order["number"] and number_counts[order["number"].casefold()] > 1:
            order["anomalies"].append({
                "id": f"duplicate-number-{order['externalId']}",
                "severity": "Advertencia",
                "type": "numero_orden_repetido",
                "description": f"El número histórico {order['number']} aparece en varias órdenes.",
                "recommendedAction": "Conservar el identificador externo y revisar antes de editar.",
                "sourceRow": order["sourceRowStart"],
            })

    detail_count = sum(len(order["details"]) for order in orders)
    if len(orders) != 280 or detail_count != 1367:
        raise SystemExit(f"Unexpected workbook totals: {len(orders)} orders, {detail_count} details")
    payload = {
        "version": "control-sales-2026-v1",
        "source": source.name,
        "orderCount": len(orders),
        "detailCount": detail_count,
        "orders": orders,
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"Wrote {len(orders)} orders and {detail_count} details to {output}")


if __name__ == "__main__":
    main()
